"""Step 6: read every sheet of the workbook, including README.

Runs FIRST despite its number. See run_ingest.py for why: step 3 must stamp
contract documents with real account ids, which only exist once this step has
run. The filenames keep the build spec's numbering so the node map in section 5
still resolves.
"""

from __future__ import annotations

import pathlib
from typing import Any

import openpyxl

from app.errors import IngestionError

EXPECTED_SHEETS = {"README", "accounts", "orders", "tickets"}


def load_workbook(path: pathlib.Path) -> dict[str, list[dict[str, Any]]]:
    if not path.exists():
        raise IngestionError(f"Workbook not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    missing = EXPECTED_SHEETS - set(wb.sheetnames)
    if missing:
        raise IngestionError(
            f"Workbook missing expected sheets: {sorted(missing)}. "
            f"Found: {wb.sheetnames}"
        )

    out: dict[str, list[dict[str, Any]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out[name] = []
            continue

        if name == "README":
            # Key/value layout rather than a header row.
            out[name] = [
                {"key": r[0], "value": r[1] if len(r) > 1 else None}
                for r in rows
                if r and r[0] is not None
            ]
            continue

        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        records = []
        for raw in rows[1:]:
            if all(c is None for c in raw):
                continue
            records.append({header[i]: raw[i] for i in range(len(header))})
        out[name] = records

    return out


def extract_snapshot_raw(readme_rows: list[dict[str, Any]]) -> str:
    """Pull the dataset snapshot string out of the README sheet.

    This single value is the reference time for every time-based calculation in
    the system, so a missing one is fatal rather than defaulted.
    """
    for row in readme_rows:
        key = str(row.get("key") or "").strip().lower()
        if "snapshot" in key:
            value = row.get("value")
            if value is None:
                break
            return str(value).strip()
    raise IngestionError(
        "No 'Dataset snapshot' row found in the README sheet. Every time-based "
        "answer depends on it; refusing to fall back to the wall clock."
    )
